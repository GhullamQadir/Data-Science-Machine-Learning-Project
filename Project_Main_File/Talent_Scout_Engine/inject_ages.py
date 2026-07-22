import os
import re
import pandas as pd
from openpyxl import load_workbook

def main():
    txt_path = r"..\..\Data Set\Age Of Players.txt"
    excel_path = r"data\Raw_data\Raw Dataset For DS & ML Project.xlsx"
    out_path = r"data\Raw_data\Raw Dataset For DS & ML Project.xlsx"

    print("Reading text file...")
    with open(txt_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse names and ages using regex
    # Matches "* Player Name = Age" or "Player Name = Age"
    # Name can have spaces and hyphens
    age_map = {}
    pattern = re.compile(r'\*?\s*([a-zA-Z\s\-\']+)\s*=\s*(\d+)')
    for match in pattern.finditer(content):
        name = match.group(1).strip()
        age = int(match.group(2))
        clean_name = re.sub(r'\s+', '', name).lower()
        age_map[clean_name] = age
    
    print(f"Parsed {len(age_map)} player ages.")

    print(f"Loading Excel file: {excel_path} (This might take a minute...)")
    # Load using openpyxl directly so we preserve formulas, styles etc if possible
    # Wait, pandas read/write is easier but drops formatting. 
    # Let's use pandas and pd.ExcelWriter to just rewrite the file safely.
    
    # Actually, openpyxl preserves it better.
    wb = load_workbook(excel_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find header row and Player column
        header_row_idx = None
        player_col_idx = None
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            for c_idx, cell_value in enumerate(row, 1):
                if str(cell_value).strip().lower() == 'player':
                    header_row_idx = r_idx
                    player_col_idx = c_idx
                    break
            if header_row_idx:
                break
                
        if not header_row_idx:
            continue
            
        # Check if Age column already exists next to Player
        age_col_idx = None
        for c_idx in range(1, ws.max_column + 2):
            cell_val = ws.cell(row=header_row_idx, column=c_idx).value
            if str(cell_val).strip().lower() == 'age':
                age_col_idx = c_idx
                break
                
        if not age_col_idx:
            # Insert Age column right after Player column
            age_col_idx = player_col_idx + 1
            ws.insert_cols(age_col_idx)
            ws.cell(row=header_row_idx, column=age_col_idx, value="Age")
            
        # Fill ages
        for r_idx in range(header_row_idx + 1, ws.max_row + 1):
            player_cell = ws.cell(row=r_idx, column=player_col_idx).value
            if player_cell:
                player_name = str(player_cell).strip().lower()
                if player_name in age_map:
                    ws.cell(row=r_idx, column=age_col_idx, value=age_map[player_name])
                    
    print("Saving updated Excel file...")
    wb.save(out_path)
    print("Done! The Ages have been successfully injected into the master dataset.")

if __name__ == "__main__":
    main()
